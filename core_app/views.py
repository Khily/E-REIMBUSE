from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model, authenticate, login, logout
from django.contrib import messages
from rest_framework import viewsets, status, serializers
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from django.db.models import Sum, Q
from datetime import datetime
from django.utils import timezone
from rest_framework_simplejwt.views import TokenObtainPairView
from collections import defaultdict
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import dari proyek Anda
from .models import Role, UserProfile, ReimbursementRequest
from .serializers import (
    CustomTokenObtainPairSerializer, RoleSerializer, UserProfileSerializer, 
    ReimbursementRequestSerializer, UserSerializer
)
# Pastikan semua form di-import
from .forms import (
    CustomUserCreationForm, CustomUserUpdateForm, UserProfileForm, RoleForm,
    ReimbursementRequestForm
)

User = get_user_model()

# --- BAGIAN API (VIEWSET) - TETAP SAMA ---
class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer

class IsAdminOrReadOnly(IsAuthenticated):
    def has_permission(self, request, view):
        return bool(
            request.user and request.user.is_authenticated and \
            (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.is_admin_role()))
            or request.method in ['GET', 'HEAD', 'OPTIONS']
        )

class IsEmployeeOrDriver(IsAuthenticated):
    def has_permission(self, request, view):
        return bool(
            request.user and request.user.is_authenticated and hasattr(request.user, 'profile') and \
            (request.user.profile.is_employee_role() or request.user.profile.is_driver())
        )

class IsManagerOrAdmin(IsAuthenticated):
    def has_permission(self, request, view):
        return bool(
            request.user and request.user.is_authenticated and hasattr(request.user, 'profile') and \
            (request.user.profile.is_supervisor_role() or request.user.profile.is_admin_role())
        )

class RoleViewSet(viewsets.ModelViewSet):
    queryset = Role.objects.all()
    serializer_class = RoleSerializer
    permission_classes = [IsAdminOrReadOnly]

class UserProfileViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = UserProfile.objects.all()
    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if not hasattr(user, 'profile'):
            return UserProfile.objects.none()
        if user.is_superuser or user.profile.is_admin_role():
            return UserProfile.objects.all()
        return UserProfile.objects.filter(user=user)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def dashboard_info(self, request):
        user = request.user
        if not hasattr(user, 'profile'):
            return Response({"detail": "User profile not found. Please contact admin."}, status=status.HTTP_404_NOT_FOUND)
        user_profile = user.profile
        role_limit = user_profile.role.monthly_limit if user_profile.role and user_profile.role.monthly_limit is not None else 0
        current_month = timezone.now().month
        current_year = timezone.now().year
        total_reimbursed_this_month = ReimbursementRequest.objects.filter(
            user=user,
            submission_date__month=current_month,
            submission_date__year=current_year,
            status__in=['Pending', 'Approved', 'Verified', 'Paid']
        ).aggregate(Sum('total_reimbursement_amount'))['total_reimbursement_amount__sum'] or 0
        remaining_limit = role_limit - total_reimbursed_this_month
        data = {
            'full_name': user.full_name,
            'personal_number': user.personal_number,
            'role_name': user_profile.role.name if user_profile.role else 'N/A',
            'plat_number': user_profile.plat_number,
            'total_monthly_limit': role_limit,
            'total_reimbursed_this_month': total_reimbursed_this_month,
            'remaining_monthly_limit': remaining_limit,
        }
        return Response(data)

class ReimbursementRequestViewSet(viewsets.ModelViewSet):
    queryset = ReimbursementRequest.objects.all()
    serializer_class = ReimbursementRequestSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if not hasattr(user, 'profile'):
            return ReimbursementRequest.objects.none()

        if user.is_superuser or user.profile.is_admin_role():
            filter_params = self.request.query_params
            date_filter = filter_params.get('date')
            month_filter = filter_params.get('month')
            year_filter = filter_params.get('year')
            status_filter = filter_params.get('status')
            queryset = ReimbursementRequest.objects.all()
            if date_filter:
                try:
                    date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
                    queryset = queryset.filter(reimburse_date__date=date_obj)
                except ValueError: pass
            if month_filter:
                try:
                    queryset = queryset.filter(reimburse_date__month=int(month_filter))
                except ValueError: pass
            if year_filter:
                try:
                    queryset = queryset.filter(reimburse_date__year=int(year_filter))
                except ValueError: pass
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            return queryset.order_by('-submission_date')
        elif user.profile.is_supervisor_role():
            return ReimbursementRequest.objects.filter(status='Pending').order_by('-submission_date')
        else:
            return ReimbursementRequest.objects.filter(user=user).order_by('-submission_date')

    def perform_create(self, serializer):
        user = self.request.user
        if not hasattr(user, 'profile'):
            raise serializers.ValidationError({"detail": "Pengguna harus memiliki profil yang lengkap untuk mengajukan reimbursement. Hubungi admin."})
        user_profile = user.profile
        if not user_profile.role:
                raise serializers.ValidationError({"detail": "Pengguna belum memiliki role. Hubungi admin."})
        role_limit = user_profile.role.monthly_limit if user_profile.role.monthly_limit is not None else 0
        current_month = timezone.now().month
        current_year = timezone.now().year
        total_reimbursed_this_month = ReimbursementRequest.objects.filter(
            user=user,
            submission_date__month=current_month,
            submission_date__year=current_year,
            status__in=['Pending', 'Approved', 'Verified', 'Paid']
        ).aggregate(Sum('total_reimbursement_amount'))['total_reimbursement_amount__sum'] or 0
        fuel_amount_liter = self.request.data.get('fuel_amount_liter')
        fuel_price = self.request.data.get('fuel_price')
        requested_amount = 0
        if fuel_amount_liter is not None and fuel_price is not None:
            try:
                requested_amount = float(fuel_amount_liter) * float(fuel_price)
            except (TypeError, ValueError):
                raise serializers.ValidationError({"detail": "Jumlah liter BBM dan harga BBM per liter harus berupa angka yang valid."})
        else:
            raise serializers.ValidationError({"detail": "Jumlah liter BBM dan harga BBM per liter wajib diisi."})
        if role_limit > 0 and (total_reimbursed_this_month + requested_amount) > role_limit:
            raise serializers.ValidationError({"detail": "Pengajuan melebihi batas reimbursement bulanan Anda."})
        owner_user_id = self.request.data.get('owner_user')
        owner_user_instance = None
        if user_profile.is_driver():
            if not owner_user_id:
                raise serializers.ValidationError({"owner_user": "Driver harus memilih pemilik kendaraan."})
            try:
                owner_user_instance = User.objects.get(id=owner_user_id)
            except User.DoesNotExist:
                raise serializers.ValidationError({"owner_user": "Pemilik kendaraan tidak ditemukan."})
        elif owner_user_id:
            raise serializers.ValidationError({"owner_user": "Karyawan tidak boleh memilih pemilik kendaraan."})
        serializer.save(user=user, owner_user=owner_user_instance)

    @action(detail=True, methods=['post'], permission_classes=[IsManagerOrAdmin])
    def approve(self, request, pk=None):
        reimbursement = self.get_object()
        if reimbursement.status == 'Pending':
            reimbursement.status = 'Approved'
            reimbursement.approved_by = request.user
            reimbursement.approved_date = timezone.now()
            reimbursement.save()
            return Response({'status': 'Pengajuan disetujui'}, status=status.HTTP_200_OK)
        return Response({'detail': 'Pengajuan tidak dapat disetujui dalam status saat ini.'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'], permission_classes=[IsManagerOrAdmin])
    def reject(self, request, pk=None):
        reimbursement = self.get_object()
        if reimbursement.status in ['Pending', 'Approved']:
            reimbursement.status = 'Rejected'
            reimbursement.approved_by = request.user
            reimbursement.approved_date = timezone.now()
            reimbursement.notes = request.data.get('notes', reimbursement.notes)
            reimbursement.save()
            return Response({'status': 'Pengajuan ditolak'}, status=status.HTTP_200_OK)
        return Response({'detail': 'Pengajuan tidak dapat ditolak dalam status saat ini.'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], permission_classes=[IsManagerOrAdmin])
    def return_for_revision(self, request, pk=None):
        reimbursement = self.get_object()
        if reimbursement.status == 'Pending':
            reimbursement.status = 'Returned'
            reimbursement.notes = request.data.get('notes', 'Dikembalikan untuk revisi.')
            reimbursement.save()
            return Response({'status': 'Pengajuan dikembalikan untuk revisi'}, status=status.HTTP_200_OK)
        return Response({'detail': 'Pengajuan tidak dapat dikembalikan dalam status saat ini.'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], permission_classes=[IsManagerOrAdmin])
    def verify(self, request, pk=None):
        reimbursement = self.get_object()
        if reimbursement.status == 'Approved':
            reimbursement.status = 'Verified'
            reimbursement.verified_by = request.user
            reimbursement.verified_date = timezone.now()
            reimbursement.save()
            return Response({'status': 'Pengajuan diverifikasi'}, status=status.HTTP_200_OK)
        return Response({'detail': 'Pengajuan belum disetujui atau sudah diverifikasi.'}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], permission_classes=[IsManagerOrAdmin])
    def mark_as_paid(self, request, pk=None):
        reimbursement = self.get_object()
        if reimbursement.status == 'Verified':
            reimbursement.status = 'Paid'
            reimbursement.save()
            return Response({'status': 'Pengajuan ditandai sudah cair'}, status=status.HTTP_200_OK)
        return Response({'detail': 'Pengajuan belum diverifikasi.'}, status=status.HTTP_400_BAD_REQUEST)

class OwnerUserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.filter(profile__plat_number__isnull=False)
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if not hasattr(user, 'profile'):
            return User.objects.none()
        if user.is_superuser or user.profile.is_admin_role() or user.profile.is_driver():
            return User.objects.filter(profile__plat_number__isnull=False)
        return User.objects.none()

# ====================================================================
# VIEWS UNTUK HALAMAN WEB (BUKAN API)
# ====================================================================

# --- Login, Logout, Halaman Admin Utama ---
@login_required
def admin_page_view(request):
    return render(request, 'admin.html', {})

# ====================================================================
# FUNGSI LOGIN YANG SUDAH DIPERBAIKI
# ====================================================================
def login_process_view(request):
    if request.method == 'POST':
        personal_number = request.POST.get('personal_number')
        password = request.POST.get('password')
        user = authenticate(request, username=personal_number, password=password)

        if user is not None:
            login(request, user)
            
            # Pengecekan role yang benar dan aman:
            # 1. Cek dulu apakah pengguna punya profil DAN sudah punya role
            if hasattr(user, 'profile') and user.profile.role is not None:
                # 2. Jika punya, baru cek apakah role-nya adalah admin
                if user.profile.is_admin_role():
                    return redirect('core_app:admin_dashboard')
            
            # 3. Jika salah satu kondisi di atas tidak terpenuhi, anggap sebagai user biasa
            return redirect('core_app:user_dashboard')
        
        else:
            messages.error(request, 'Personal Number atau Password salah.')
            # Arahkan kembali ke halaman login utama (ganti 'home' jika perlu)
            return redirect('home')
            
    return redirect('home')

def logout_view(request):
    logout(request)
    return redirect('home')

# --- User Management (CRUD) ---
@login_required
def manage_user_view(request):
    all_users = User.objects.filter(is_superuser=False)
    
    users_data_with_limit = []
    current_month = timezone.now().month
    current_year = timezone.now().year

    for user in all_users:
        try:
            profile = user.profile
        except UserProfile.DoesNotExist:
            profile = None

        role_limit = 0
        if profile and profile.role and profile.role.monthly_limit is not None:
            role_limit = profile.role.monthly_limit

        total_reimbursed = ReimbursementRequest.objects.filter(
            Q(user=user) | Q(owner_user=user),
            submission_date__month=current_month,
            submission_date__year=current_year,
            status__in=['Pending', 'Approved', 'Verified', 'Paid']
        ).distinct().aggregate(total=Sum('total_reimbursement_amount'))['total'] or 0
        
        remaining_limit = role_limit - total_reimbursed
        
        users_data_with_limit.append({
            'user': user,
            'profile': profile,
            'remaining_limit': remaining_limit
        })

    context = {
        'users_data': users_data_with_limit
    }
    return render(request, 'manage-user.html', context)

@login_required
def add_user_view(request):
    if request.method == 'POST':
        user_form = CustomUserCreationForm(request.POST)
        profile_form = UserProfileForm(request.POST)
        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save()
            profile = profile_form.save(commit=False)
            profile.user = user
            profile.save()
            messages.success(request, 'Pengguna baru berhasil ditambahkan.')
            return redirect('core_app:manage_user')
    else:
        user_form = CustomUserCreationForm()
        profile_form = UserProfileForm()
    context = {
        'user_form': user_form,
        'profile_form': profile_form
    }
    return render(request, 'add-user.html', context)

@login_required
def edit_user_view(request, pk):
    user_to_edit = get_object_or_404(User, pk=pk)
    profile, created = UserProfile.objects.get_or_create(user=user_to_edit)

    if request.method == 'POST':
        user_form = CustomUserUpdateForm(request.POST, instance=user_to_edit)
        profile_form = UserProfileForm(request.POST, instance=profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, 'Data pengguna berhasil diperbarui.')
            return redirect('core_app:manage_user')
    else:
        user_form = CustomUserUpdateForm(instance=user_to_edit)
        profile_form = UserProfileForm(instance=profile)
    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'profile': profile
    }
    return render(request, 'edit-user.html', context)

@login_required
def delete_user_view(request, pk):
    user_to_delete = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user_to_delete.delete()
        messages.success(request, 'Pengguna berhasil dihapus.')
        return redirect('core_app:manage_user')
    return redirect('core_app:manage_user')


# --- Role Management (CRUD) ---
@login_required
def manage_role_view(request):
    all_roles = Role.objects.all()
    context = {'roles': all_roles}
    return render(request, 'manage-role.html', context)

@login_required
def add_role_view(request):
    if request.method == 'POST':
        form = RoleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Role baru berhasil ditambahkan.')
            return redirect('core_app:manage_role')
    else:
        form = RoleForm()
    context = {'form': form}
    return render(request, 'add-role.html', context)

@login_required
def edit_role_view(request, pk):
    role = get_object_or_404(Role, pk=pk)
    if request.method == 'POST':
        form = RoleForm(request.POST, instance=role)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data role berhasil diperbarui.')
            return redirect('core_app:manage_role')
    else:
        form = RoleForm(instance=role)
    context = {
        'form': form,
        'role': role
    }
    return render(request, 'edit-role.html', context)

@login_required
def delete_role_view(request, pk):
    role = get_object_or_404(Role, pk=pk)
    if request.method == 'POST':
        role.delete()
        messages.success(request, 'Pengguna berhasil dihapus.')
        return redirect('core_app:manage_role')
    return redirect('core_app:manage_role')


# --- Views Lainnya ---
@login_required
def view_request_view(request):
    # Hanya admin yang bisa melihat halaman ini
    if not hasattr(request.user, 'profile') or not request.user.profile.is_admin_role():
        messages.error(request, 'Anda tidak memiliki hak akses.')
        return redirect('core_app:user_dashboard')
    
    # Ambil semua requests sebagai base query
    all_requests = ReimbursementRequest.objects.all()
    
    # Filter berdasarkan bulan jika ada (menggunakan reimburse_date)
    selected_month = request.GET.get('month')
    if selected_month:
        try:
            month_int = int(selected_month)
            all_requests = all_requests.filter(reimburse_date__month=month_int)
        except (ValueError, TypeError):
            pass
    
    # Filter berdasarkan tahun jika ada (menggunakan reimburse_date)
    selected_year = request.GET.get('year')
    if selected_year:
        try:
            year_int = int(selected_year)
            all_requests = all_requests.filter(reimburse_date__year=year_int)
        except (ValueError, TypeError):
            pass
    
    # Urutkan berdasarkan tanggal reimburse terbaru
    all_requests = all_requests.order_by('-reimburse_date')
    
    context = {
        'all_requests': all_requests,
        'selected_month': selected_month,
        'selected_year': selected_year,
    }
    return render(request, 'view-request.html', context)

@login_required
def user_dashboard_view(request):
    context = {}
    user = request.user
    if hasattr(user, 'profile'):
        user_profile = user.profile
        role_limit = user_profile.role.monthly_limit if user_profile.role and user_profile.role.monthly_limit is not None else 0
        current_month = timezone.now().month
        current_year = timezone.now().year
        
        reimbursements_this_month = ReimbursementRequest.objects.filter(
            Q(user=user) | Q(owner_user=user),
            submission_date__month=current_month,
            submission_date__year=current_year,
            status__in=['Pending', 'Approved', 'Verified', 'Paid']
        ).distinct()
        
        total_reimbursed = reimbursements_this_month.aggregate(total=Sum('total_reimbursement_amount'))['total'] or 0
        total_liters = reimbursements_this_month.aggregate(total=Sum('fuel_amount_liter'))['total'] or 0

        remaining_limit = role_limit - total_reimbursed
        progress_percentage = 0
        if role_limit > 0:
            progress_percentage = (total_reimbursed / role_limit) * 100

        context = {
            'user': user,
            'profile': user_profile,
            'total_limit': role_limit,
            'total_reimbursed': total_reimbursed,
            'remaining_limit': remaining_limit,
            'total_liters': total_liters,
            'progress_percentage': progress_percentage,
        }
    return render(request, 'user-dashboard.html', context)

# --- Views untuk Reimbursement Form ---
@login_required
def reimburse_form_view(request):
    user = request.user
    if not hasattr(user, 'profile') or not user.profile.role:
        messages.error(request, "Profil atau role Anda belum lengkap. Hubungi admin.")
        return redirect('core_app:user_dashboard')

    user_profile = user.profile
    role_limit = user_profile.role.monthly_limit if user_profile.role.monthly_limit is not None else 0
    current_month = timezone.now().month
    current_year = timezone.now().year
    total_reimbursed = ReimbursementRequest.objects.filter(
        Q(user=user) | Q(owner_user=user),
        submission_date__month=current_month,
        submission_date__year=current_year,
        status__in=['Pending', 'Approved', 'Verified', 'Paid']
    ).distinct().aggregate(total=Sum('total_reimbursement_amount'))['total'] or 0
    remaining_limit = role_limit - total_reimbursed
    limit_exceeded = remaining_limit <= 0 and role_limit > 0

    if request.method == 'POST':
        if limit_exceeded:
            messages.error(request, 'Anda sudah mencapai limit reimbursement bulan ini.')
            return redirect('core_app:reimburse_form')

        form = ReimbursementRequestForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            new_amount = form.cleaned_data['fuel_amount_liter'] * form.cleaned_data['fuel_price']
            if (total_reimbursed + new_amount) > role_limit and role_limit > 0:
                messages.error(request, f'Pengajuan melebihi sisa limit Anda (Rp {remaining_limit:,.0f}).')
                context = {'form': form, 'limit_exceeded': False, 'remaining_limit': remaining_limit}
                return render(request, 'reimburse-form.html', context)

            reimbursement = form.save(commit=False)
            reimbursement.user = request.user
            reimbursement.save()
            messages.success(request, 'Pengajuan reimburse berhasil dikirim!')
            return redirect('core_app:reimburse_success', pk=reimbursement.pk)
    else:
        form = ReimbursementRequestForm(user=request.user)

    context = {
        'form': form,
        'limit_exceeded': limit_exceeded,
        'remaining_limit': remaining_limit
    }
    return render(request, 'reimburse-form.html', context)

@login_required
def reimburse_success_view(request, pk):
    reimbursement = get_object_or_404(ReimbursementRequest, pk=pk)
    context = {'reimbursement': reimbursement}
    return render(request, 'reimburse-success.html', context)

# --- View untuk Halaman Riwayat ---
@login_required
def history_view(request):
    user = request.user
    
    # Ambil filter dari URL
    selected_month = request.GET.get('month')
    selected_year = request.GET.get('year')

    # Query dasar
    requests_list = ReimbursementRequest.objects.filter(
        Q(user=user) | Q(owner_user=user)
    ).distinct()

    # Terapkan filter jika ada
    if selected_month and selected_month.isdigit():
        requests_list = requests_list.filter(submission_date__month=selected_month)
    if selected_year and selected_year.isdigit():
        requests_list = requests_list.filter(submission_date__year=selected_year)
    
    requests_list = requests_list.order_by('-submission_date')
    
    # Ambil data untuk dropdown
    all_dates = ReimbursementRequest.objects.filter(Q(user=user) | Q(owner_user=user)).dates('submission_date', 'month', order='DESC')
    available_years = sorted(list(set([d.year for d in all_dates])))
    months = [
        (1, "Januari"), (2, "Februari"), (3, "Maret"), (4, "April"),
        (5, "Mei"), (6, "Juni"), (7, "Juli"), (8, "Agustus"),
        (9, "September"), (10, "Oktober"), (11, "November"), (12, "Desember")
    ]
    
    # Kelompokkan hasil yang sudah difilter
    history_groups = defaultdict(list)
    for req in requests_list:
        history_groups[req.submission_date.date()].append(req)
        
    context = {
        'history_groups': dict(history_groups),
        'available_years': available_years,
        'months': months,
        'selected_month': int(selected_month) if selected_month and selected_month.isdigit() else None,
        'selected_year': int(selected_year) if selected_year and selected_year.isdigit() else None,
    }
    return render(request, 'history.html', context)

# --- View untuk Panduan ---
@login_required
def panduan_view(request):
    return render(request, 'panduan.html', {})

# --- Views untuk Aksi Admin (Approve/Reject) ---
@login_required
def approve_request_view(request, pk):
    reimbursement = get_object_or_404(ReimbursementRequest, pk=pk)
    if request.method == 'POST':
        if reimbursement.status == 'Pending':
            reimbursement.status = 'Approved'
            reimbursement.approved_by = request.user
            reimbursement.approved_date = timezone.now()
            reimbursement.save()
            messages.success(request, 'Pengajuan berhasil disetujui.')
        else:
            messages.warning(request, 'Pengajuan ini tidak dapat disetujui.')
    return redirect('core_app:view_request')

@login_required
def reject_request_view(request, pk):
    reimbursement = get_object_or_404(ReimbursementRequest, pk=pk)
    if request.method == 'POST':
        if reimbursement.status == 'Pending':
            reimbursement.status = 'Rejected'
            reimbursement.approved_by = request.user
            reimbursement.approved_date = timezone.now()
            reimbursement.save()
            messages.success(request, 'Pengajuan berhasil ditolak.')
        else:
            messages.warning(request, 'Pengajuan ini tidak dapat ditolak.')
    return redirect('core_app:view_request')

# --- View untuk Halaman Detail/Evidence ---
@login_required
def evidence_view(request, pk):
    # Cek apakah pengguna punya profil dan merupakan admin
    if not hasattr(request.user, 'profile') or not request.user.profile.is_admin_role():
        # Jika bukan admin, kirim pesan error dan kembalikan ke dashboard user
        messages.error(request, 'Anda tidak memiliki hak akses untuk melihat halaman ini.')
        return redirect('core_app:user_dashboard')

    # Jika lolos pengecekan (adalah admin), lanjutkan seperti biasa
    reimbursement = get_object_or_404(ReimbursementRequest, pk=pk)
    context = {
        'req': reimbursement
    }
    return render(request, 'evidence.html', context)

@login_required
def export_requests_excel(request):
    # Hanya admin yang bisa mengakses
    if not hasattr(request.user, 'profile') or not request.user.profile.is_admin_role():
        messages.error(request, 'Anda tidak memiliki hak akses.')
        return redirect('core_app:user_dashboard')
    
    # Ambil semua requests dengan filter yang sama seperti view_request_view
    all_requests = ReimbursementRequest.objects.all()
    
    # Filter berdasarkan bulan jika ada (menggunakan reimburse_date)
    selected_month = request.GET.get('month')
    if selected_month:
        try:
            month_int = int(selected_month)
            all_requests = all_requests.filter(reimburse_date__month=month_int)
        except (ValueError, TypeError):
            pass
    
    # Filter berdasarkan tahun jika ada (menggunakan reimburse_date)
    selected_year = request.GET.get('year')
    if selected_year:
        try:
            year_int = int(selected_year)
            all_requests = all_requests.filter(reimburse_date__year=year_int)
        except (ValueError, TypeError):
            pass
    
    # Urutkan berdasarkan tanggal reimburse terbaru
    all_requests = all_requests.order_by('-reimburse_date')
    
    # Buat workbook dan worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Reimbursement"
    
    # Style untuk header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header columns
    headers = [
        'No', 'Tanggal Reimburse', 'Tanggal Pengajuan', 'Nama Pengaju', 'Plat Nomor', 
        'Jumlah Reimbursement', 'Status', 'Tanggal Persetujuan', 'Disetujui Oleh'
    ]
    
    # Tulis header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Tulis data
    for row_num, request in enumerate(all_requests, 2):
        # No
        cell = ws.cell(row=row_num, column=1)
        cell.value = row_num - 1
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        
        # Tanggal Reimburse
        cell = ws.cell(row=row_num, column=2)
        cell.value = request.reimburse_date.strftime("%d/%m/%Y") if request.reimburse_date else "-"
        cell.border = border
        
        # Tanggal Pengajuan
        cell = ws.cell(row=row_num, column=3)
        cell.value = request.submission_date.strftime("%d/%m/%Y %H:%M")
        cell.border = border
        
        # Nama Pengaju
        cell = ws.cell(row=row_num, column=4)
        cell.value = request.user.full_name if hasattr(request.user, 'full_name') else request.user.username
        cell.border = border
        
        # Plat Nomor
        cell = ws.cell(row=row_num, column=5)
        cell.value = request.get_plat_number
        cell.border = border
        
        # Jumlah Reimbursement
        cell = ws.cell(row=row_num, column=6)
        cell.value = f"Rp {request.total_reimbursement_amount:,.0f}"
        cell.border = border
        
        # Status
        cell = ws.cell(row=row_num, column=7)
        cell.value = request.status
        cell.border = border
        # Warna berdasarkan status
        if request.status == 'Approved':
            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        elif request.status == 'Rejected':
            cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        elif request.status == 'Pending':
            cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        
        # Tanggal Persetujuan
        cell = ws.cell(row=row_num, column=8)
        cell.value = request.approved_date.strftime("%d/%m/%Y %H:%M") if request.approved_date else "-"
        cell.border = border
        
        # Disetujui Oleh
        cell = ws.cell(row=row_num, column=9)
        if request.approved_by:
            cell.value = request.approved_by.full_name if hasattr(request.approved_by, 'full_name') else request.approved_by.username
        else:
            cell.value = "-"
        cell.border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Buat response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename berdasarkan filter
    filename = "Laporan_Reimbursement"
    if selected_month or selected_year:
        filename += "_Filter"
        if selected_month:
            month_names = {
                '1': 'Januari', '2': 'Februari', '3': 'Maret', '4': 'April',
                '5': 'Mei', '6': 'Juni', '7': 'Juli', '8': 'Agustus',
                '9': 'September', '10': 'Oktober', '11': 'November', '12': 'Desember'
            }
            filename += f"_{month_names.get(selected_month, selected_month)}"
        if selected_year:
            filename += f"_{selected_year}"
    
    filename += f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook ke response
    wb.save(response)
    
    return response